#!/usr/bin/env python3
import argparse
from openpyxl import load_workbook


def add_links(path: str):
    wb = load_workbook(path)
    if 'Overview' not in wb.sheetnames:
        return False
    ws = wb['Overview']

    # Place hyperlinks in row 0-based index +1 (Excel row 1)
    # E1: Controls, F1: Parameters
    cell_controls = 'E1'
    ws[cell_controls] = 'Go to Controls'
    ws[cell_controls].hyperlink = "#'Controls'!A1"
    ws[cell_controls].style = 'Hyperlink'

    cell_params = 'F1'
    ws[cell_params] = 'Go to Parameters'
    ws[cell_params].hyperlink = "#'Parameters'!A1"
    ws[cell_params].style = 'Hyperlink'

    wb.save(path)
    return True


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--dashboard-path', required=True)
    args = ap.parse_args()
    ok = add_links(args.dashboard_path)
    if not ok:
        print('Overview sheet not found; no links added')


if __name__ == '__main__':
    main()
